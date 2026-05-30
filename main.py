import streamlit as st
import pandas as pd
from datetime import datetime
import tempfile
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

st.set_page_config(page_title="АПП Генератор ПВЗ", layout="wide")
st.title("🛠 Генератор АПП для ПВЗ")

# Инициализация
if 'scanned' not in st.session_state:
    st.session_state.scanned = []
if 'df' not in st.session_state:
    st.session_state.df = None
if 'fio' not in st.session_state:
    st.session_state.fio = ""

# ====================== БОКОВАЯ ПАНЕЛЬ ======================
with st.sidebar:
    st.header("Смена")
    fio = st.text_input("Ваше ФИО", value=st.session_state.fio, placeholder="Фамилия Имя Отчество")
    if fio:
        st.session_state.fio = fio

    st.divider()
    uploaded_file = st.file_uploader("1. Загрузить CSV", type=['csv'])
    
    if uploaded_file:
        try:
            sample = uploaded_file.read(2048).decode('utf-8', errors='ignore')
            uploaded_file.seek(0)
            
            if 'sep=' in sample.lower():
                df = pd.read_csv(uploaded_file, sep=None, engine='python', encoding='utf-8', skiprows=1)
            else:
                try:
                    df = pd.read_csv(uploaded_file, sep=',', encoding='utf-8')
                except:
                    df = pd.read_csv(uploaded_file, sep=';', encoding='utf-8')
            
            df.columns = [str(col).strip().replace('"', '').replace("'", "") for col in df.columns]
            st.session_state.df = df
            st.success(f"✅ Загружено {len(df)} строк")
            st.write("Колонки:", list(df.columns))
        except Exception as e:
            st.error(f"❌ Ошибка чтения файла: {e}")

# ====================== СКАНИРОВАНИЕ ======================
col1, col2 = st.columns([3, 1])

with col1:
    st.subheader("2. Сканирование заказов")
    with st.form(key="scan_form", clear_on_submit=True):
        order_input = st.text_input("Номер заказа", placeholder="260597447132 или LM-260597447132")
        submitted = st.form_submit_button("✅ Добавить", type="primary", use_container_width=True)
        
        if submitted and order_input.strip():
            if st.session_state.df is None:
                st.error("Сначала загрузите CSV файл!")
            else:
                raw = str(order_input).strip()
                search_num = raw.replace("LM-", "").replace("lm-", "").strip()
                
                df = st.session_state.df
                found = False
                order_columns = [col for col in df.columns if any(w in col.lower() for w in ['заказ', 'order', 'номер'])]
                if not order_columns:
                    order_columns = df.columns

                for col in order_columns:
                    df[col] = df[col].astype(str).str.strip()
                    result = df[df[col] == search_num]
                    if not result.empty:
                        row = result.iloc[0].to_dict()
                        if row not in st.session_state.scanned:
                            st.session_state.scanned.append(row)
                            st.success(f"✅ УСПЕШНО ДОБАВЛЕН: {search_num}")
                        else:
                            st.error("🔴 ДУБЛИКАТ!")
                        found = True
                        break
                if not found:
                    st.error(f"❌ Заказ {search_num} не найден")

with col2:
    st.subheader("Статистика")
    st.metric("Отсканировано", len(st.session_state.scanned))

# ====================== НЕОТСКАНИРОВАННЫЕ ======================
if st.session_state.df is not None:
    st.subheader("📋 Неотсканированные заказы")
    df = st.session_state.df.copy()
    if "Статус заказа" in df.columns and "Статус контроля" in df.columns:
        mask = (
            df["Статус заказа"].astype(str).str.contains('Собран', na=False) &
            df["Статус контроля"].astype(str).str.contains('пройден', na=False)
        )
        remaining = df[mask].copy()
        
        if st.session_state.scanned:
            scanned_set = {str(r.get('Заказ', '')) for r in st.session_state.scanned}
            remaining = remaining[~remaining['Заказ'].astype(str).isin(scanned_set)]
        
        st.write(f"**Осталось отсканировать: {len(remaining)}**")
        if not remaining.empty:
            st.dataframe(remaining, use_container_width=True, hide_index=True)

# ====================== ОТСКАНИРОВАННЫЕ ======================
st.subheader(f"✅ Отсканированные ({len(st.session_state.scanned)})")
if st.session_state.scanned:
    st.dataframe(pd.DataFrame(st.session_state.scanned), use_container_width=True, hide_index=True)

# ====================== ФОРМИРОВАНИЕ АПП ======================
if st.button("🚀 Сформировать все АПП", type="secondary", use_container_width=True):
    if not st.session_state.scanned:
        st.error("Нет отсканированных заказов")
    else:
        scanned_df = pd.DataFrame(st.session_state.scanned)
        current_fio = st.session_state.fio or "______________________"
        current_date = datetime.now().strftime("%d.%m.%Y")
        
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            path = tmp.name
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            partner_names = {
                '135_FIVEPOST': '5Post',
                '135_SDEK': 'SDEK',
                '135_Yandex': 'Yandex',
                'UNKNOWN': 'DPD'
            }
            
            # ==================== ШАПКИ ====================
            headers = {
                '135_FIVEPOST': [
                    'ООО "ЛЕ МОНЛИД" (магазин по адресу Лемана про Московское шоссе 14 лит, А), в лице Директора направления доставки до клиента в омниканальном формате Мамонова Дмитрия Сергеевича,  действующего на основании Доверенности № 00/21/62 от 08.04.2021 г., именуемый(-ый) в дальнейшем Заказчик, и',
                    'ООО "ФАЙВ ПОСТ" в лице Халиповского Кирилла Алексеевича, действующего на основании доверенности, удостоверенной нотариусом г. Москвы Алексеева С.Д., номер в реестре: №77/649-н/77-2023-1-21 от 11.01.2023 года, именуемое в дальнейшем Исполнитель,',
                    'настоящим актом удостоверяют, что Заказчик передал, а Исполнитель принял в рамках исполнения договора  № 00/19089 от 27.08.2024 г. следующие отправление для последующей доставки Получателям:'
                ],
                '135_SDEK': [
                    'ООО "ЛЕ МОНЛИД" (магазин по адресу Лемана про Московское шоссе 14 лит, А), в лице Директора по управлению цепями поставок Эрика Пуле,  действующего на основании Доверенности № 00/1054 от 22.12.2014 г,, именуемый(-ый) в дальнейшем Клиент, и',
                    'ООО СДЭК-Глобал в лице Исполнительного директора Островского Артема Феликсовича, действующего на основании Доверенности №3 от 09.01.2017 г,, именуемое(-ый) в дальнейшем Исполнитель,',
                    'настоящим актом удостоверяет, что Клиент передал, а Исполнитель принял в рамках исполнения договора № 00/15557 от 31.07.2023 следующие отправление для последующей доставки Получателям:'
                ],
                '135_Yandex': [
                    'ООО "ЛЕ МОНЛИД" (магазин по адресу Лемана про Московское шоссе 14 лит, А), в лице Директора направления доставки до клиента в омниканальном формате Мамонова Дмитрия Сергеевича,  действующего на основании Доверенности № 00/21/62 от 08,04,2021 г,, именуемый(ый) в дальнейшем Заказчик, и',
                    'ООО "Яндекс Доставка" в лице Генерального директора Р,А, Морозова, действующего на основании Устава, именуемое в дальнейшем "Яндекс",',
                    'настоящим актом удостоверяют, что Заказчик передал, а Яндекс принял в рамках исполнения договора  № 01/780 от 01,11,2021 г, следующие отправление для последующей доставки Получателям:'
                ],
                'UNKNOWN': [
                    'ООО "ЛЕ МОНЛИД" (магазин по адресу Лемана про Московское шоссе 14 лит, А), в лице Директора направления доставки до клиента в омниканальном формате Мамонова Дмитрия Сергеевича,  действующего на основании Доверенности № 00/21/62 от 08.04.2021 г., именуемый(-ый) в дальнейшем Заказчик, и',
                    'АО "ДПД РУС" в лице Генерального директора Воинова Н.Ю, действующего на основании Уставаа, именуемое в дальнейшем Исполнитель,',
                    'настоящим актом удостоверяют, что Заказчик передал, а Исполнитель принял в рамках исполнения договора  № 00/5224 от 23.04.2018 г. следующие отправление для последующей доставки Получателям:'
                ]
            }
            
            for code, sheet_name in partner_names.items():
                group = scanned_df[scanned_df.get('Партнер', '') == code]
                if group.empty:
                    continue
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Шапка
                ws['A1'] = "Акт приема-передачи"
                ws.merge_cells('A1:D1')
                ws['A1'].font = Font(bold=True, size=14)
                ws['A1'].alignment = Alignment(horizontal="center")
                
                ws['A2'] = f"от {current_date}"
                ws.merge_cells('A2:D2')
                ws['A2'].alignment = Alignment(horizontal="center")
                
                row = 4
                for line in headers.get(code, headers['UNKNOWN']):
                    ws[f'A{row}'] = line
                    ws.merge_cells(f'A{row}:D{row}')
                    ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical="top")
                    row += 1
                
                # Таблица
                table_start = row + 2
                for col, text in enumerate(["№", "номер отправления", "вес отправления (кг)", "стоимость отправления(руб,)"], 1):
                    cell = ws.cell(row=table_start, column=col, value=text)
                    cell.font = Font(bold=True)
                
                total_weight = 0.0
                total_cost = 0
                
                for i, r in enumerate(group.itertuples(index=False), 1):
                    row_dict = dict(zip(group.columns, r))
                    curr_row = table_start + i
                    ws.cell(row=curr_row, column=1, value=i)
                    ws.cell(row=curr_row, column=2, value=str(row_dict.get('Заказ', '')))
                    w = float(row_dict.get('Вес', 0))
                    c = float(row_dict.get('Цена заказа', 0))
                    ws.cell(row=curr_row, column=3, value=round(w, 3))
                    ws.cell(row=curr_row, column=4, value=int(c))
                    total_weight += w
                    total_cost += c
                
                # Итого
                last = table_start + len(group)
                ws.cell(row=last+1, column=1, value="Итого:").font = Font(bold=True)
                ws.cell(row=last+1, column=2, value=len(group))
                ws.cell(row=last+1, column=3, value=round(total_weight, 3))
                ws.cell(row=last+1, column=4, value=int(total_cost))
            
            wb.save(path)
            
            with open(path, 'rb') as f:
                st.download_button(
                    label="📥 Скачать АПП (все партнёры в одном файле)",
                    data=f.read(),
                    file_name=f"АПП_Все_{current_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            st.success("✅ АПП успешно сформирован!")
            
        except Exception as e:
            st.error(f"Ошибка: {e}")
        finally:
            try:
                if os.path.exists(path):
                    os.unlink(path)
            except:
                pass

st.caption("Дата и шапки на месте • Один файл")







































































































































































































































































































































































































































































































































































